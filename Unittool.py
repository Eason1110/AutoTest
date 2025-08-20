import tkinter as tk
from tkinter import filedialog, ttk
import re
import os
import sys
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string # 引入 column_index_from_string
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.worksheet.cell_range import CellRange
from bs4 import BeautifulSoup
import locale
import subprocess
from datetime import datetime

'''
update: 2025/08/20
'''

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PY Files & Test Report Tool")
        self.geometry("900x700")
        
        self.py_files = {} 
        self.selected_cases_by_file = {} 
        self.tree_file_nodes = {} 
        self.last_py_folder = None 
        self.last_html_report_folder = None
        self.testplan_files_in_result = [] 
        self.last_excel_save_path = None
        self.process = None
        
        self.status_label = tk.Label(self, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W, font=("Arial", 10))
        self.status_label.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill="both", padx=10, pady=10)

        self.py_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.py_tab, text="PY File Handling")
        self.create_py_tab_widgets(self.py_tab)

        self.excel_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.excel_tab, text="Excel Report Handling")
        self.create_excel_tab_widgets(self.excel_tab)
        
    def show_status_message(self, message, message_type="info"):
        color = "black" 
        if message_type == "warning":
            color = "orange"
        elif message_type == "error":
            color = "red"
        elif message_type == "success":
            color = "green"
        
        self.status_label.config(text=message, fg=color)

    def create_py_tab_widgets(self, parent_frame):
        file_frame = tk.LabelFrame(parent_frame, text="Load PY Files", padx=10, pady=10)
        file_frame.pack(fill="x", padx=10, pady=5)

        self.file_list_label = tk.Label(file_frame, text="No PY files loaded")
        self.file_list_label.pack(side=tk.LEFT, expand=True, fill="x")

        select_files_btn = tk.Button(file_frame, text="Load PY Files", command=self.load_py_files)
        select_files_btn.pack(side=tk.RIGHT)

        option_frame = tk.Frame(parent_frame, padx=10, pady=5)
        option_frame.pack(fill="x", padx=10, pady=5)

        select_all_btn = tk.Button(option_frame, text="Select All Test Items", command=self.select_all_test_items)
        select_all_btn.pack(side=tk.LEFT, padx=5)

        deselect_all_btn = tk.Button(option_frame, text="Deselect All Test Items", command=self.deselect_all_test_items)
        deselect_all_btn.pack(side=tk.LEFT, padx=5)

        self.selected_count_label = tk.Label(option_frame, text="Selected items: 0")
        self.selected_count_label.pack(side=tk.LEFT, padx=10)
        
        self.open_py_folder_btn = tk.Button(option_frame, text="Open PY Folder", command=self.open_last_py_folder, state=tk.DISABLED)
        self.open_py_folder_btn.pack(side=tk.RIGHT, padx=5)

        export_btn = tk.Button(option_frame, text="Export Unittest Plan", command=self.export_unittest_plan)
        export_btn.pack(side=tk.RIGHT)

        search_frame = tk.LabelFrame(parent_frame, text="Test Item Analysis", padx=10, pady=10)
        search_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(search_frame, text='Note: This will analyze all "test_case" items inside PY files. Follow naming conventions.', fg="blue").pack(side=tk.LEFT, padx=5)

        result_frame = tk.LabelFrame(parent_frame, text="Test Item Selection Results", padx=10, pady=10)
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(result_frame, columns=("No.", "Item Name", "Select"), show="headings")
        self.tree.heading("No.", text="No.", anchor=tk.CENTER)
        self.tree.heading("Item Name", text="Item Name", anchor=tk.W)
        self.tree.heading("Select", text="Select", anchor=tk.CENTER)

        self.tree.column("No.", width=50, anchor=tk.CENTER)
        self.tree.column("Item Name", width=400, anchor=tk.W)
        self.tree.column("Select", width=70, anchor=tk.CENTER)

        self.tree.tag_configure("file_node", background="#D3D3D3", foreground="blue", font=("", 9, "bold"))

        self.tree.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.tree.config(yscrollcommand=scrollbar.set)

        self.tree.bind("<ButtonRelease-1>", self.on_tree_click) 
        
    def open_last_py_folder(self):
        if self.last_py_folder and os.path.isdir(self.last_py_folder):
            try:
                if os.sys.platform == "win32":
                    os.startfile(self.last_py_folder)
                elif os.sys.platform == "darwin": 
                    subprocess.Popen(["open", self.last_py_folder])
                else: 
                    subprocess.Popen(["xdg-open", self.last_py_folder])
                self.show_status_message(f"Opened PY folder: {self.last_py_folder}", "info")
            except Exception as e:
                self.show_status_message(f"Cannot open PY folder: {e}", "error")
        else:
            self.show_status_message("No PY folder path available. Please load PY files first.", "warning")

    def create_excel_tab_widgets(self, parent_frame):
        excel_settings_frame = tk.LabelFrame(parent_frame, text="Excel Settings", padx=10, pady=10)
        excel_settings_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(excel_settings_frame, text="Read Column (Col A, Row B):").grid(row=0, column=0, sticky="w", pady=2)
        self.read_col_entry = tk.Entry(excel_settings_frame, width=5)
        self.read_col_entry.grid(row=0, column=1, padx=2, pady=2)
        self.read_row_entry = tk.Entry(excel_settings_frame, width=5)
        self.read_row_entry.grid(row=0, column=2, padx=2, pady=2)
        self.read_col_entry.insert(0, "A") 
        self.read_row_entry.insert(0, "1") 

        tk.Label(excel_settings_frame, text="Write Column (Col A, Row B):").grid(row=1, column=0, sticky="w", pady=2)
        self.write_col_entry = tk.Entry(excel_settings_frame, width=5)
        self.write_col_entry.grid(row=1, column=1, padx=2, pady=2)
        self.write_row_entry = tk.Entry(excel_settings_frame, width=5)
        self.write_row_entry.grid(row=1, column=2, padx=2, pady=2)
        self.write_col_entry.insert(0, "B") 
        self.write_row_entry.insert(0, "1") 

        excel_buttons_frame = tk.Frame(parent_frame, padx=10, pady=5)
        excel_buttons_frame.pack(fill="x", padx=10, pady=5)

        load_testplan_btn = tk.Button(excel_buttons_frame, text="Load Testplan (Excel)", command=self.load_testplan)
        load_testplan_btn.pack(side=tk.LEFT, padx=5)

        write_results_btn = tk.Button(excel_buttons_frame, text="Write Results (HTML -> Excel)", command=self.write_results_to_excel)
        write_results_btn.pack(side=tk.LEFT, padx=5)

        self.open_report_folder_btn = tk.Button(excel_buttons_frame, text="Open Report Folder", command=self.open_last_report_folder, state=tk.DISABLED)
        self.open_report_folder_btn.pack(side=tk.RIGHT, padx=5)

    def open_last_report_folder(self):
        target_folder = None
        if self.last_excel_save_path and os.path.isdir(os.path.dirname(self.last_excel_save_path)):
            target_folder = os.path.dirname(self.last_excel_save_path)
            message_type = "Excel Results"
        elif self.last_html_report_folder and os.path.isdir(self.last_html_report_folder):
            target_folder = self.last_html_report_folder
            message_type = "HTML Report"
        else:
            self.show_status_message("No report folder path available. Please load an HTML report or save Excel results first.", "warning")
            return

        try:
            if os.sys.platform == "win32":
                os.startfile(target_folder)
            elif os.sys.platform == "darwin": 
                subprocess.Popen(["open", target_folder])
            else: 
                subprocess.Popen(["xdg-open", target_folder])
            self.show_status_message(f"Opened {message_type} folder: {target_folder}", "info")
        except Exception as e:
            self.show_status_message(f"Cannot open report folder: {e}", "error")


    def load_py_files(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Python files", "*.py")])
        if file_paths:
            self.py_files.clear()
            self.last_py_folder = os.path.dirname(file_paths[0])

            for file_path in file_paths:
                module_name = os.path.splitext(os.path.basename(file_path))[0]
                self.py_files[file_path] = {'module_name': module_name, 'test_class_name': None, 'report_dir': None}

            if len(self.py_files) > 0:
                displayed_names = ", ".join([os.path.basename(path) for path in self.py_files.keys()])
                self.file_list_label.config(text=f"Selected: {displayed_names}")
                self.analyze_all_py_files()
                self.open_py_folder_btn.config(state=tk.NORMAL)
                self.show_status_message(f"Loaded {len(file_paths)} PY files.", "success")
            else:
                self.file_list_label.config(text="No PY files loaded")
                self.open_py_folder_btn.config(state=tk.DISABLED)
                self.show_status_message("No PY files loaded.", "warning")
        else:
            self.show_status_message("Cancelled loading PY files.", "info")

    def analyze_all_py_files(self):
        if not self.py_files:
            self.show_status_message("Please load PY files first!", "warning")
            return

        self.selected_cases_by_file.clear()
        self.tree.delete(*self.tree.get_children())
        self.tree_file_nodes.clear()

        total_cases_found = 0
        overall_no = 1

        # Regex to find output path
        report_output_pattern = re.compile(r"output=['\"](.*?)['\"]")

        test_case_pattern = re.compile(r'def\s+(test_case[a-zA-Z0-9_]+)\s*\(self\):')
        test_class_pattern = re.compile(r'class\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*\(\s*unittest\.TestCase\s*\):')

        for py_file_path, file_info in self.py_files.items():
            file_display_name = os.path.basename(py_file_path)
            file_node_id = py_file_path

            self.tree.insert("", "end", iid=file_node_id,
                            values=("", file_display_name, "☐"),
                            tags=("file_node", "checkbox"))
            self.tree_file_nodes[py_file_path] = file_node_id

            self.tree.item(file_node_id, open=True)

            self.selected_cases_by_file[py_file_path] = {}

            try:
                with open(py_file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                output_match = report_output_pattern.search(content)
                if output_match:
                    file_info['report_dir'] = output_match.group(1)
                    print(f"Found report path in file {file_display_name}: {file_info['report_dir']}")
                else:
                    file_info['report_dir'] = 'D:/SeleniumProject/test_reports'  # default path

                class_matches = list(test_class_pattern.finditer(content))
                if class_matches:
                    detected_class_name = class_matches[0].group(1)
                    file_info['test_class_name'] = detected_class_name
                    self.tree.item(file_node_id, values=("", f"{file_display_name} (Class: {detected_class_name})", "☐"))
                else:
                    file_info['test_class_name'] = "MyTestCase"
                    self.tree.item(file_node_id, values=("", f"{file_display_name} (No test class detected, default MyTestCase)", "☐"))

                matches = test_case_pattern.finditer(content)

                file_cases_count = 0
                for match in matches:
                    case_name = match.group(1)
                    if case_name not in self.selected_cases_by_file[py_file_path]:
                        var = tk.BooleanVar(value=False)
                        self.selected_cases_by_file[py_file_path][case_name] = var

                        child_iid = f"{file_node_id}::{case_name}"
                        self.tree.insert(file_node_id, "end", iid=child_iid,
                                        values=(overall_no, case_name, "☐"), tags=("checkbox",))

                        total_cases_found += 1
                        file_cases_count += 1
                        overall_no += 1

                if file_cases_count == 0:
                    current_values = list(self.tree.item(file_node_id, "values"))
                    if "(Class:" in current_values[1]:
                        current_values[1] = current_values[1].replace(")", ", no test cases)")
                    else:
                        current_values[1] = f"{os.path.basename(py_file_path)} (no test cases)"
                    self.tree.item(file_node_id, values=current_values)

            except Exception as e:
                self.tree.item(file_node_id, values=("", f"{file_display_name} (Failed to read: {e})", "☐"))
                self.show_status_message(f"Error reading or parsing file '{file_display_name}': {e}", "error")

        self.update_selected_count_label()
        if total_cases_found > 0:
            self.show_status_message(f"Successfully analyzed all PY files, found {total_cases_found} test cases.", "success")
        else:
            self.show_status_message("No test cases found in loaded PY files.", "warning")

    def on_tree_click(self, event):
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if not item_id:
            return

        if column == "#3": 
            tags = self.tree.item(item_id, "tags")
            
            if "file_node" in tags:
                py_file_path = item_id 
                
                all_children_iids = self.tree.get_children(item_id)
                if not all_children_iids: 
                    return

                any_unselected = False
                for child_iid in all_children_iids:
                    parts = child_iid.split('::', 1)
                    if len(parts) == 2:
                        child_py_path = parts[0]
                        child_case_name = parts[1]
                        if child_py_path == py_file_path: 
                            if child_case_name in self.selected_cases_by_file[child_py_path]:
                                if not self.selected_cases_by_file[child_py_path][child_case_name].get():
                                    any_unselected = True
                                    break
                
                new_state = any_unselected 

                for child_iid in all_children_iids:
                    parts = child_iid.split('::', 1)
                    if len(parts) == 2:
                        child_py_path = parts[0]
                        child_case_name = parts[1] 

                        if child_py_path == py_file_path:
                            if child_case_name in self.selected_cases_by_file[child_py_path]:
                                self.selected_cases_by_file[child_py_path][child_case_name].set(new_state)
                                self.update_checkbox_display(child_iid, new_state)
                
                self.update_file_node_checkbox_display(py_file_path)
                self.update_selected_count_label()

            else: 
                parts = item_id.split('::', 1)
                if len(parts) < 2: 
                    return 

                py_file_path = parts[0]
                case_name = parts[1] 

                if py_file_path in self.selected_cases_by_file and \
                   case_name in self.selected_cases_by_file[py_file_path]:
                    current_var = self.selected_cases_by_file[py_file_path][case_name]
                    current_var.set(not current_var.get()) 
                    self.update_checkbox_display(item_id, current_var.get())
                    self.update_file_node_checkbox_display(py_file_path) 
                    self.update_selected_count_label()

    def update_checkbox_display(self, item_id, is_selected):
        current_values = list(self.tree.item(item_id, "values"))
        current_values[2] = "☑" if is_selected else "☐"
        self.tree.item(item_id, values=current_values)

    def update_file_node_checkbox_display(self, py_file_path):
        if py_file_path not in self.tree_file_nodes:
            return

        file_node_id = self.tree_file_nodes[py_file_path]
        file_cases = self.selected_cases_by_file.get(py_file_path, {})
        
        if not file_cases: 
            current_values = list(self.tree.item(file_node_id, "values"))
            current_values[2] = "☐" 
            self.tree.item(file_node_id, values=current_values)
            return

        total_cases = len(file_cases)
        selected_cases = sum(1 for var in file_cases.values() if var.get())

        current_values = list(self.tree.item(file_node_id, "values"))
        if selected_cases == total_cases:
            current_values[2] = "☑" 
        elif selected_cases > 0:
            current_values[2] = "■" 
        else:
            current_values[2] = "☐" 

        self.tree.item(file_node_id, values=current_values)

    def update_selected_count_label(self):
        count = 0
        for file_cases in self.selected_cases_by_file.values():
            count += sum(1 for var in file_cases.values() if var.get())
        self.selected_count_label.config(text=f"Selected test cases: {count}") 

    def select_all_test_items(self):
        for py_file_path, file_cases in self.selected_cases_by_file.items():
            for case_name, var in file_cases.items():
                var.set(True)
                item_id = f"{py_file_path}::{case_name}" 
                self.update_checkbox_display(item_id, True)
            self.update_file_node_checkbox_display(py_file_path) 
        self.update_selected_count_label()
        self.show_status_message("All test cases selected.", "info")

    def deselect_all_test_items(self):
        for py_file_path, file_cases in self.selected_cases_by_file.items():
            for case_name, var in file_cases.items():
                var.set(False)
                item_id = f"{py_file_path}::{case_name}" 
                self.update_checkbox_display(item_id, False)
            self.update_file_node_checkbox_display(py_file_path) 
        self.update_selected_count_label()
        self.show_status_message("All test cases deselected.", "info")


    def export_unittest_plan(self):
        selected_cases_by_module = {}

        for py_file_path, file_cases in self.selected_cases_by_file.items():
            file_info = self.py_files[py_file_path]
            module_name = os.path.splitext(os.path.basename(py_file_path))[0]
            test_class_name = file_info['test_class_name'] if file_info['test_class_name'] else "MyTestCase"
            report_dir = file_info['report_dir'] if file_info.get('report_dir') else 'D:/SeleniumProject/test_reports'

            for case_name, var in file_cases.items():
                if var.get():
                    if module_name not in selected_cases_by_module:
                        selected_cases_by_module[module_name] = {'class_name': test_class_name, 'cases': [], 'report_dir': report_dir}
                    selected_cases_by_module[module_name]['cases'].append(case_name)

        if not selected_cases_by_module:
            self.show_status_message("Please select at least one test case!", "warning")
            return

        initial_dir = self.last_py_folder if self.last_py_folder and os.path.isdir(self.last_py_folder) else None

        file_path = filedialog.asksaveasfilename(
            defaultextension=".py",
            filetypes=[("Python files", "*.py"), ("All files", "*.*")],
            initialfile="Unittest_plan.py",
            initialdir=initial_dir
        )

        if not file_path:
            self.show_status_message("Unittest Plan export canceled.", "info")
            return

        output_content = [
            "import unittest\n",
            "import os\n",
            "import sys\n",
            "from datetime import datetime\n",
            "import locale\n",
            "\n"
        ]

        output_content.append("if getattr(sys, 'frozen', False):  # exe environment\n")
        output_content.append("    current_script_dir = sys._MEIPASS\n")
        output_content.append("else:\n")
        output_content.append("    current_script_dir = os.path.dirname(os.path.abspath(__file__))\n")
        output_content.append("\n")
        output_content.append("     html_test_runner_dir = os.path.join(current_script_dir, '_internal')\n")
        output_content.append("     sys.path.insert(0, html_test_runner_dir)\n")
        output_content.append("\n")
        output_content.append("try:\n")
        output_content.append("    import HTMLTestRunner\n")
        output_content.append("except ImportError as e:\n")
        output_content.append("    print('Error: Unable to import HTMLTestRunner module')\n")
        output_content.append("    print(f'Please make sure HTMLTestRunner is located in {html_test_runner_dir} folder.')\n")
        output_content.append("    raise e\n\n")

        for mod_name, info in sorted(selected_cases_by_module.items()):
            output_content.append(f"from {mod_name} import {info['class_name']}\n")

        output_content.append("\nif __name__ == '__main__':\n")

        for mod_name, info in sorted(selected_cases_by_module.items()):
            current_class_name = info['class_name']
            current_report_dir = info['report_dir']

            if not info['cases']:
                continue

            output_content.append(f"    print(f\"\\n--- Running tests for {mod_name}.py ---\")\n")
            output_content.append(f"    suite = unittest.TestSuite()\n")
            output_content.append(f"    class_name = '{current_class_name}'\n")

            selected_test_cases_str = ', '.join([f"'{case_name}'" for case_name in info['cases']])
            output_content.append(f"    selected_cases_list = [{selected_test_cases_str}]\n")
            output_content.append(f"    cases_description = 'Includes test cases: ' + ', '.join(selected_cases_list)\n")

            for case_name in info['cases']:
                output_content.append(f"    suite.addTest({current_class_name}('{case_name}'))\n")
            output_content.append(f"    report_path = os.path.join(r'{current_report_dir}', "
                                f"f'{current_class_name}_{{datetime.now().strftime(\"%Y%m%d_%H%M%S\")}}.html')\n")
            output_content.append(f"    with open(report_path, 'w', encoding='utf-8') as f:\n")
            output_content.append(f"        runner = HTMLTestRunner.HTMLTestRunner(\n")
            output_content.append(f"            verbosity=2,\n")
            output_content.append(f"            title=f'Test Report for {current_class_name} ({mod_name})',\n")
            output_content.append(f"        )\n")
            output_content.append(f"        runner.run(suite)\n")
            output_content.append(f"    print(f'Test report saved to: {{report_path}}')\n\n")

        output_content.append("    print(\"\\n--- All selected test suites have been executed and reported. ---\")\n")

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.writelines(output_content)
            self.show_status_message(f"Unittest Plan successfully exported to '{os.path.basename(file_path)}' in 'D:/SeleniumProject/test_reports'. You can open the PY folder.", "success")
        except Exception as e:
            self.show_status_message(f"Error occurred while exporting file: {e}", "error")


    def load_testplan(self):
        excel_file_paths = filedialog.askopenfilenames(
            title="Select Testplan (Excel) files",
            filetypes=[
                ("Excel files (XLSX)", "*.xlsx"), 
                ("Excel files (XLS)", "*.xls")
            ]
        )
        if not excel_file_paths:
            self.show_status_message("Excel Testplan load canceled.", "info")
            return

        result_dir = "Result"
        os.makedirs(result_dir, exist_ok=True)
        
        self.testplan_files_in_result.clear()
        
        for f_path in excel_file_paths:
            dest_path = os.path.join(result_dir, os.path.basename(f_path))
            try:
                shutil.copy(f_path, dest_path)
                self.testplan_files_in_result.append(dest_path)
                self.show_status_message(f"Successfully loaded and copied Testplan: {os.path.basename(f_path)}", "success")
            except Exception as e:
                self.show_status_message(f"Error copying file '{os.path.basename(f_path)}': {e}", "error")

        if self.testplan_files_in_result:
            self.show_status_message(f"Successfully loaded {len(self.testplan_files_in_result)} Testplan files to Result folder.", "success")
            self.open_report_folder_btn.config(state=tk.NORMAL)
        else:
            self.show_status_message("No Excel Testplan files selected.", "warning")


    def write_results_to_excel(self):
        html_dir = filedialog.askdirectory(title="Select folder containing HTML reports")
        if not html_dir:
            self.show_status_message("HTML report folder selection canceled.", "info")
            return
        
        self.last_html_report_folder = html_dir 
        self.open_report_folder_btn.config(state=tk.NORMAL) 

        excel_files_to_process = self.testplan_files_in_result
        
        if not excel_files_to_process:
            self.show_status_message("No Excel Testplan files found in Result folder! Please load first.", "warning")
            return

        try:
            read_col_str = self.read_col_entry.get().upper()
            read_row_int = int(self.read_row_entry.get()) - 1 
            write_col_str = self.write_col_entry.get().upper()
            write_row_int = int(self.write_row_entry.get()) - 1

            if not read_col_str.isalpha() or not write_col_str.isalpha():
                self.show_status_message("Read/Write columns must be letters (e.g., A, B)!", "error")
                return
            if read_row_int < 0 or write_row_int < 0:
                self.show_status_message("Read/Write rows must be positive integers!", "error")
                return

            all_html_results = {}
            html_files = [f for f in os.listdir(html_dir) if f.endswith(".html")]
            if not html_files:
                self.show_status_message("No HTML report files found in the selected folder.", "warning")
                return
            
            for html_file in html_files:
                file_path = os.path.join(html_dir, html_file)
                content = None
                encodings_to_try = ['utf-8', 'gbk', 'cp950', 'latin-1'] 
                for encoding in encodings_to_try:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            content = f.read()
                        break 
                    except UnicodeDecodeError:
                        pass
                    except Exception as e:
                        self.show_status_message(f"Error reading HTML file '{html_file}': {e}", "warning")
                        break

                if content is None:
                    self.show_status_message(f"Unable to read HTML file '{html_file}'.", "warning")
                    continue
                
                try:
                    soup = BeautifulSoup(content, 'html.parser')
                    results = self.parse_html_report(soup)
                    for item in results:
                        all_html_results[item['name']] = item['result'] 
                except Exception as e:
                    self.show_status_message(f"Error parsing HTML file '{html_file}': {e}", "warning")
            
            if not all_html_results:
                self.show_status_message("No results extracted from HTML reports.", "warning")
                return

            # Process each Testplan
            total_results_written = 0
            for testplan_path_in_result in excel_files_to_process:
                self.show_status_message(f"Processing Testplan: {os.path.basename(testplan_path_in_result)}...", "info")
                try:
                    workbook = load_workbook(testplan_path_in_result)
                    sheet = workbook.active
                    results_written_count = 0
                    
                    for row_idx, row_data in enumerate(sheet.iter_rows()):
                        current_excel_row_num = row_idx + 1
                        if current_excel_row_num < read_row_int + 1:
                            continue

                        testcase_name_cell = sheet[f"{read_col_str}{current_excel_row_num}"]
                        testcase_name = str(testcase_name_cell.value).strip() if testcase_name_cell.value else ""

                        if testcase_name in all_html_results:
                            result_to_write = all_html_results[testcase_name]
                            write_cell = sheet[f"{write_col_str}{current_excel_row_num}"]
                            write_cell.value = result_to_write

                            # Default style
                            write_cell.font = Font()
                            write_cell.border = Border()
                            write_cell.fill = PatternFill(fill_type=None)
                            write_cell.alignment = Alignment()
                            
                            # Apply style based on result
                            if result_to_write.lower() == "pass":
                                write_cell.font = Font(color="008000", bold=True)
                            elif result_to_write.lower() == "fail":
                                write_cell.font = Font(color="FF0000", bold=True)
                                write_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            
                            results_written_count += 1

                    # Save file with suffix to avoid overwriting
                    base, ext = os.path.splitext(os.path.basename(testplan_path_in_result))
                    current_time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    new_filename = f"{base}_Result_{current_time_str}{ext}"
                    new_save_path = os.path.join(os.path.dirname(testplan_path_in_result), new_filename)
                    
                    workbook.save(new_save_path)
                    self.show_status_message(f"Testplan '{os.path.basename(testplan_path_in_result)}' processed, results saved as '{new_filename}'.", "success")
                    total_results_written += results_written_count

                except Exception as e:
                    self.show_status_message(f"Error processing file '{os.path.basename(testplan_path_in_result)}': {e}", "error")

            self.show_status_message(f"All Testplan batch processing completed. Total results written: {total_results_written}.", "success")
            self.last_excel_save_path = new_save_path

        except Exception as e:
            self.show_status_message(f"Unexpected error during batch processing: {e}", "error")


    def parse_html_report(self, soup):
        all_test_results = []
        for tr_tag in soup.find_all('tr', class_=['hiddenRow', 'none']):
            test_info = {}

            name_tag_container = tr_tag.find('td', class_=['passCase', 'failCase'])
            if name_tag_container:
                name_tag = name_tag_container.find('div', class_='testcase').find('a', class_='popup_link')
                if name_tag:
                    test_info['name'] = name_tag.get_text(strip=True)
                else:
                    test_info['name'] = "N/A (Name link not found)"
            else:
                test_info['name'] = "N/A (Name cell not found)"

            result_tag_container = tr_tag.find('td', align='center')
            if result_tag_container:
                result_link = result_tag_container.find('button').find('a', class_='popup_link')
                if result_link:
                    test_info['result'] = result_link.get_text(strip=True)
                else:
                    test_info['result'] = "N/A (Result link not found)"
            else:
                test_info['result'] = "N/A (Result cell not found)"

            if test_info['name'].startswith("test_case") and test_info['result'] not in ["N/A (Result link not found)", "N/A (Result cell not found)"]:
                all_test_results.append(test_info)
        return all_test_results

if __name__ == "__main__":
    app = App()
    app.mainloop()